from tkinter import *
from tkinter import messagebox
from datetime import date
import mysql.connector as sql
from tkinter import filedialog
from PIL import Image ,ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib
from tkcalendar import Calendar,DateEntry

background="#00999C"
framebg="#EDEDED"
framefg="#06283D"

root=Tk()
root.title("Student Dashboard")
root.geometry("1250x720+210+100")
root.config(bg=background)
root.resizable(False,False)

#icon image
image_icon=PhotoImage(file="Image/icon.png")
root.iconphoto(False,image_icon)

file=pathlib.Path('Student_record.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration no."
    sheet['B1']="Name"
    sheet['C1']="Address"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of Registration"
    sheet['G1']="Email Addresss"
    sheet['H1']="Phone Number"
    sheet['I1']="Father Name"
    sheet['J1']="Father's Occupation"
    sheet['K1']="Mother Name"
    sheet['L1']="Mother's Occupation"

    file.save('Student_record.xlsx')



#genders
def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
        # print(gender)
    else:
        gender="Female"
        # print(gender)


#Exit Window
def Exit():
    root.destroy()
#Upload image
def show_image():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),title="Select image file",filetype=(("JPG File","*.jpg"),("PNG File","*.png"),("All files","*.txt")))

    img=(Image.open(filename))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

#####################  Registration NO #######################
#since we have  to fill manually registrartion number each time 
#here is the code for  automatic registration number entry system
def registration_NO():
    file=openpyxl.load_workbook("Student_record.xlsx")
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value
    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")

##################   Clear  #################
def Clear():
    global img
    Name.set("")
    # cal._select("")
    Address.set("")
    Email.set("")
    Phone.set("")
    F_name.set("")
    F_occupation.set("")
    M_name.set("")
    M_occupation.set("")
    cal._set_text("SELECT")

    registration_NO()
    saveButton.config(state='normal')
    img1=PhotoImage(file="Image/upload photo.png")
    lbl.config(image=img1)
    lbl.image=img1
    img=""

#################SAVE###############
def save():
    R1=Registration.get()
    D1=Date.get()
    N1=Name.get()
    A1=Address.get()
    try:
        G1=gender #if we haven't selected a gender then it will show error message
    except:
        messagebox.showerror("Error","Select Gender!")

    D2=cal.get_date()
    D2_1= D2.strftime("%m/%d/%Y")
    E1=Email.get()
    P1=Phone.get()
    fn=F_name.get()
    mn=M_name.get()
    F1=F_occupation.get()
    M1=M_occupation.get()
    # print(R1)
    # print(D1)
    # print(N1)
    # print(A1)
    # print(G1)
    # print(D2_1)
    # print(E1)
    # print(P1)
    # print(fn)
    # print(mn)
    # print(F1)
    # print(M1)

    if N1=="" or A1=="" or E1=="" or P1=="" or D2=="SELECT" or fn=="" or mn=="" or F1=="" or M1=="":
        messagebox.showerror("Error","Few Data is missing !")
    else:
        file=openpyxl.load_workbook("Student_record.xlsx")
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=A1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=D2_1)
        sheet.cell(column=6,row=sheet.max_row,value=D1)
        sheet.cell(column=7,row=sheet.max_row,value=E1)
        sheet.cell(column=8,row=sheet.max_row,value=P1)
        sheet.cell(column=9,row=sheet.max_row,value=fn)
        sheet.cell(column=10,row=sheet.max_row,value=F1)
        sheet.cell(column=11,row=sheet.max_row,value=mn)
        sheet.cell(column=12,row=sheet.max_row,value=M1)

        file.save(r'Student_record.xlsx')

        try:
            img.save("Studnet Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("Info","Profile Picture is not available !!")
        messagebox.showinfo("Info","Successfully data created !!")
        Clear() # clear entry box and image section
        registration_NO() #it will recheck no and reissue new no.
        import Fee_Frontend


def search():
    text=Search.get() #taking input from entry box
    Clear() # to clear all the available data present in window
    saveButton.config(state='disabled')# save btn disable so that no can click on it
    file=openpyxl.load_workbook("Student_record.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value ==int(text):
            name=row[0]
            # print(str(name))
            reg_no_position=str(name)[14:-1] #gives like A2,A3,A4....An
            reg_number=str(name)[15:-1] #gives number like 2,3,4....
            # print(reg_no_position)
            # print(reg_number)
    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid registration number !!")

    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=6).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value
    # print(x1)
    # print(x2)
    # print(x3)
    # print(x4)
    # print(x5)
    # print(x6)
    # print(x7)
    # print(x8)
    # print(x9)
    # print(x10)
    # print(x11)
    # print(x12)

    Registration.set(x1)
    Name.set(x2)
    Address.set(x3)
    if x4=="Female":
        R2.select()
    else:
        R1.select()
    cal.set_date(x5)
    Date.set(x6)
    Email.set(x7)
    Phone.set(x8)
    F_name.set(x9)
    F_occupation.set(x10)
    M_name.set(x11)
    M_occupation.set(x12)

    img=(Image.open("Student Images/"+str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2


###############Update######################
def Update():
    R1=Registration.get()
    D1=Date.get()
    N1=Name.get()
    A1=Address.get()
    selection()
    G1=gender
    D2=cal.get_date()
    D2_1= D2.strftime("%m/%d/%Y")
    E1=Email.get()
    P1=Phone.get()
    fn=F_name.get()
    mn=M_name.get()
    F1=F_occupation.get()
    M1=M_occupation.get()

    file=openpyxl.load_workbook("Student_record.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value ==R1:
            name=row[0]
            print(str(name))
            reg_no_position=str(name)[14:-1] #gives like A2,A3,A4....An
            reg_number=str(name)[15:-1]
            print(reg_number)

    # sheet.cell(column=1,row=int(reg_number),value=R1) ###### since there is no need to update registration number and it is same as before
    sheet.cell(column=2,row=int(reg_number),value=N1)
    sheet.cell(column=3,row=int(reg_number),value=A1)
    sheet.cell(column=4,row=int(reg_number),value=G1)
    sheet.cell(column=5,row=int(reg_number),value=D2_1)
    sheet.cell(column=6,row=int(reg_number),value=D1)
    sheet.cell(column=7,row=int(reg_number),value=E1)
    sheet.cell(column=8,row=int(reg_number),value=P1)
    sheet.cell(column=9,row=int(reg_number),value=fn)
    sheet.cell(column=10,row=int(reg_number),value=F1)
    sheet.cell(column=11,row=int(reg_number),value=mn)
    sheet.cell(column=12,row=int(reg_number),value=M1)      

    file.save(r'Student_record.xlsx')
    try:
        img.save("Student Images/"+str(R1)+".jpg")
    except:
        pass

    messagebox.showinfo("Update","Update Successfully !!")
    Clear()


#top frames
label=Label(root,text="STUDENT MANAGEMENT SYSTEM",width=12,height=3,bg='#99999c',anchor="center",font='arial 14 bold')
label.pack(side=TOP,fill=X)
label=Label(root,text="STUDENT REGISTRATION",width=10,height=2,bg='#3498db',fg='#fff',font='arial 14 bold underline' ,anchor='center')
label.pack(side=TOP,fill=X)


#search box to update
Search=StringVar()
srch=Entry(root,textvariable=Search,width=15,bd=2,font="arial 18")
srch.place(x=880,y=76)
#search button
image_icon=PhotoImage(file="Image\search.png")
srch_btn=Button(root,text="SEARCH",compound=LEFT,image=image_icon,width=98,height=30,bg='#68ddfa',font='arial 10 bold',command=search)
srch_btn.place(x=1096,y=75)


#update button
up_pic=PhotoImage(file="Image/Layer 4.png")
update_button=Button(root,image=up_pic,bg='#3498db', width='50',height='40',command=Update)
update_button.place(x=40,y=74)

#Registration and date
Label(root,text="Registration NO:",font="arial 13",fg=framefg,bg=background).place(x=30,y=150)
Label(root,text="Date",font="arial 13",fg=framefg,bg=background).place(x=500,y=150)

Registration=IntVar()
Date=StringVar()

reg_entry=Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=150)

registration_NO() 

today=date.today()
d1=today.strftime("%d/%m/%Y")
# print(d1)
date_entry=Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)
Date.set(d1)

#student details
obj=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj,text="Address:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="E-mail ID:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Phone No:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name=StringVar()
name_entry=Entry(obj,textvariable=Name,width=20,font="arial 10",bd=2)
name_entry.place(x=160,y=50)

# # DOB=StringVar()
# # dob_entry=Entry(obj,textvariable=DOB,width=20,font="arial 10")
# # dob_entry.place(x=160,y=100)

cal=DateEntry(obj,width=20,bg=background,fg="white",bd=2,date_pattern='dd/mm/yyyy')
cal.pack(pady=20)
cal.place(x=160,y=100)
# cal.set_date("")

radio=IntVar()
R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)
R2=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=200,y=150)

Address=StringVar()
add_entry=Entry(obj,textvariable=Address,width=20,font="arial 10",bd=2)
add_entry.place(x=630,y=50)

Email=StringVar()
em_entry=Entry(obj,textvariable=Email,width=20,font="arial 10",bd=2)
em_entry.place(x=630,y=100)

Phone=StringVar()
phone_entry=Entry(obj,textvariable=Phone,width=20,font="arial 10",bd=2)
phone_entry.place(x=630,y=150)

# Class=Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12'],font="Roboto 10",width=17,state="r")
# Class.place(x=630,y=50)
# Class.set("Select Class")


#parents details
obj2=LabelFrame(root,text="Parent's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Father's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
# Label(obj2,text="Phone No:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj2,text="Mother's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
# Label(obj2,text="Phone No:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)


F_name=StringVar()
f_entry=Entry(obj2,textvariable=F_name,width=20,font="arial 10")
f_entry.place(x=160,y=50)

F_occupation=StringVar()
FO_entry=Entry(obj2,textvariable=F_occupation,width=20,font="arial 10")
FO_entry.place(x=160,y=100)

M_name=StringVar()
m_entry=Entry(obj2,textvariable=M_name,width=20,font="arial 10")
m_entry.place(x=630,y=50)

M_occupation=StringVar()
MO_entry=Entry(obj2,textvariable=M_occupation,width=20,font="arial 10")
MO_entry.place(x=630,y=100)

#image
f=Frame(root,bd=3,bg="#ffffff",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)
img=PhotoImage(file="Image/upload photo.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

#button
Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=show_image).place(x=1000,y=370)
saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=save)
saveButton.place(x=1000,y=450)
Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightyellow",command=Clear).place(x=1000,y=530)
Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="grey",command=Exit).place(x=1000,y=610)

root.mainloop()